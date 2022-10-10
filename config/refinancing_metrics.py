import pandas as pd
import openpyxl
import traceback
from datetime import datetime, timedelta

from config import api_yandex_async
from common import BaseExcel


class RefinancingExcel(BaseExcel):
    def __int__(self):
        self.filename = r"C:\Users\Центрофинанс\OneDrive - ООО Микрокредитная компания «Центрофинанс Групп»\Рабочий стол\OneDrive - ООО Микрокредитная компания «Центрофинанс Групп»\Документы\CR Перекредитование в ЛК.xlsx"
        self.book = openpyxl.load_workbook(self.filename)

    def get_ids_refinancing(self, is_comment=True):
        result = {}
        for name in self.book.sheetnames:
            sheet = self.book[name]
            result[name] = {} if is_comment else []

            for row in sheet.iter_rows(max_row=1, min_col=2, max_col=sheet.max_column):
                for cell in row:
                    if is_comment:
                        if cell.comment is not None:
                            result[name].update({cell.comment.text.split("\n")[1]: cell.column_letter})
                    else:
                        if cell.comment is None:
                            result[name].append(cell.column)

        return result

    def do_offset_formulas(self, sheet, data, row):
        for col in data:
            cell = sheet.cell(row=row, column=col)
            cell_offset = cell.offset(row=-1)
            cell_offset_row = cell_offset.row
            try:
                if type(cell_offset.value) is str:
                    if cell_offset.value[0] == "=":
                        result = cell_offset.value.replace(str(cell_offset_row), str(row))
                        self.logger.info(
                            f"{sheet}: {cell.coordinate}: {cell.value} --> {cell_offset.value.replace(str(cell_offset_row), str(row))}")
                        sheet[cell.coordinate] = result

            except Exception:
                self.logger.error(traceback.format_exc())

    def main(self):
        data = self.get_ids_refinancing()
        data_formulas = self.get_ids_refinancing(is_comment=False)

        for name, ids in data.items():
            sheet = self.book[name]
            last_row = self.get_row_for_write(sheet)

            first_date = str(last_row[0])
            row = int(last_row[1])
            yesterday = (datetime.now() - timedelta(days=1)).strftime('20%y-%m-%d')
            daterange = pd.date_range(first_date.split(" ")[0], yesterday)

            self.logger.info(f"{last_row}:{yesterday}")

            if first_date.split(" ")[0] != datetime.now().strftime('20%y-%m-%d'):
                for date in daterange:
                    date_format = str(date.strftime("%Y-%m-%d"))
                    metrics = api_yandex_async.main(ids, date1=date_format, date2=date_format, need_users=False)

                    sheet[f"A{row + 1}"] = date + timedelta(days=1)

                    for col, goal in metrics.items():
                        if goal != "" and col != "date":
                            sheet[str(col) + str(row)] = int(goal)

                    self.do_offset_formulas(sheet, data_formulas[name], row)
                    row += 1

        self.book.save(self.filename)

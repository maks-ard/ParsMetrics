import json
import platform
import openpyxl
from openpyxl.styles.numbers import BUILTIN_FORMATS
import logging
import traceback
import psutil
import pandas as pd
from datetime import datetime, timedelta
from tkinter import filedialog, Tk

from config import api_yandex_async
from config.editor import GetIdRow

logger = logging.getLogger("main")


def get_yesterday(day_or_month):
    return (datetime.now() - timedelta(days=1)).strftime('%d') if day_or_month == "day" else (
            datetime.now() - timedelta(days=1)).strftime('%m')


def start_file(filename=None):  # открытие ecxel файла
    import os

    if filename is None:
        os.startfile(file_for_write())
    else:
        os.startfile(filename)


def get_path_ecxel():  # открывает окно для выбора файла ecxel и возвращает полный путь
    root = Tk()
    root.attributes("-topmost", True)
    root.lift()
    root.withdraw()

    return filedialog.askopenfilename()


def choice_file(path_to_files):
    print("Выбери в какой файл записывать данные!")
    name_pc = platform.node()
    path = get_path_ecxel()
    path_to_files[name_pc] = path

    with open(r"data/path_to_ecxel.json", "w", encoding="utf-8") as file:
        json.dump(path_to_files, file, indent=3, ensure_ascii=False)

    return path


def file_for_write():  # выбор пути, в зависимости от устройства
    path_to_files = json.load(open(r"data/path_to_ecxel.json", encoding="utf-8"))

    try:
        for proc in psutil.process_iter():
            if proc.name() == "EXCEL.EXE":
                proc.kill()
        file = open(path_to_files[platform.node()])  # проверяет корректность пути до файла и не открыт ли он
        file.close()
        return path_to_files[platform.node()]

    except (KeyError, FileNotFoundError):
        logger.error(traceback.format_exc())
        return choice_file(path_to_files)


def name_sheet(month=None):
    if month is None:
        month = datetime.now().strftime("%m")
    return json.load(open(r"data/name_sheet.json", encoding="utf-8"))[month]  # текущий месяц


def edit_file(day=get_yesterday("day"), month=get_yesterday("month"), date1='yesterday', date2='yesterday'):
    filename = file_for_write()  # путь к ecxel файлу
    name_sheet_now = name_sheet(month)
    ids: dict = GetIdRow(filename).get_id_row("BR")

    metrics = api_yandex_async.main(ids, date1=date1, date2=date2)  # выгруженные метрики
    book = openpyxl.load_workbook(filename=filename)  # книга ecxel

    sheet = book[name_sheet_now]  # нужный лист в ecxel
    col = json.load(open(r'data/date_col.json', encoding="utf-8"))[day]  # нужная колонка

    for key, index in metrics.items():
        if key != "date" and index != '':
            sheet[col + str(key)] = int(index)  # запись значения в ячейки
    book.save(filename=filename)  # сохранение изменений


def get_ids_refin(book, is_comment=True):
    result = {}
    for name in book.sheetnames:
        sheet = book[name]
        result[name] = {} if is_comment else []

        for row in sheet.iter_rows(max_row=1, min_col=2, max_col=sheet.max_column):
            for cell in row:
                if is_comment:
                    if cell.comment is not None:
                        result[name].update({cell.comment.text.split("\n")[1]: cell.column_letter})
                else:
                    if cell.comment is None:
                        result[name].append(cell.column)

    return result


def get_row_for_write(sheet):
    rows = sheet.iter_rows(max_col=1)
    result = ()
    for row in rows:
        for cell in row:
            result = (cell.value, cell.row)
            if cell.value is None:
                return cell.offset(row=-1).value, cell.offset(row=-1).row
    return result


def do_offset_formulas(sheet, data, row):
    for col in data:
        cell = sheet.cell(row=row, column=col)
        cell_offset = cell.offset(row=-1)
        cell_offset_row = cell_offset.row
        try:
            if type(cell_offset.value) is str:
                if cell_offset.value[0] == "=":
                    sheet[cell.coordinate] = cell_offset.value.replace(str(cell_offset_row), str(row))
                    if "РСД" not in cell_offset.value:
                        sheet[cell.coordinate].number_format = BUILTIN_FORMATS[9]


        except Exception:
            logger.error(traceback.format_exc())


def write_goal_refinance():
    # filename = r"C:\Users\Центрофинанс\OneDrive - ООО Микрокредитная компания «Центрофинанс Групп»\Рабочий стол\OneDrive - ООО Микрокредитная компания «Центрофинанс Групп»\Документы\CR Перекредитование в ЛК.xlsx"
    filename = r"CR Перекредитование в ЛК.xlsx" # тестовый
    book = openpyxl.load_workbook(filename)
    data = get_ids_refin(book)
    data_formulas = get_ids_refin(book, is_comment=False)

    for name, ids in data.items():
        sheet = book[name]
        last_row = get_row_for_write(sheet)

        first_date = str(last_row[0])
        row = int(last_row[1])
        yesterday = (datetime.now() - timedelta(days=1)).strftime('20%y-%m-%d')
        daterange = pd.date_range(first_date.split(" ")[0], yesterday)

        logger.info(f"{last_row}:{yesterday}")

        if first_date.split(" ")[0] != datetime.now().strftime('20%y-%m-%d'):
            for date in daterange:
                date_format = str(date.strftime("%Y-%m-%d"))
                metrics = api_yandex_async.main(ids, date1=date_format, date2=date_format, need_users=False)

                date: datetime
                sheet[f"A{row + 1}"] = date + timedelta(days=1)
                sheet[f"A{row + 1}"].number_format = "DD.MM.YYYY"

                for col, goal in metrics.items():
                    if goal != "" and col != "date":
                        sheet[str(col) + str(row)] = int(goal)

                do_offset_formulas(sheet, data_formulas[name], row)
                row += 1

    book.save(filename)

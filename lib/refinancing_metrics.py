import pandas as pd
import traceback
from datetime import datetime

import openpyxl
import plyer
import schedule
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.comments import Comment
from openpyxl.workbook import Workbook

from common import BaseExcel, api_yandex_async
from common import YandexApi


class RefinancingExcel(BaseExcel):
    def __init__(self, filename=None):
        super().__init__()
        self.filename = self.get_filepath("CR Перекредитование в ЛК") if filename is None else self.get_filepath(
            filename)
        self.api = YandexApi()

    @staticmethod
    def get_row_for_write(sheet: Worksheet) -> tuple:
        """Получить последную дату и ряд"""
        return [(row[0].value, row[0].row) for row in sheet.iter_rows(max_col=1) if row[0].value is not None][-1]

    @staticmethod
    def get_row_by_date(sheet: Worksheet, date: datetime):
        for row in sheet.iter_rows(max_col=1):
            if row[0].value.date() == date.date():
                return row[0].value, row[0].row

    @staticmethod
    def get_ids_refinancing(book: Workbook, is_comment=True) -> dict[dict] | dict[list]:
        """
        is_comment: True - Получить id целей и их колонки dict[dict]
        is_comment: False - Получить колонки для формул dict[list]
        """
        result = {name: {} if is_comment else [] for name in book.sheetnames}

        for name in book.sheetnames:
            sheet: Worksheet = book[name]

            for row in sheet.iter_rows(max_row=1, min_col=2, max_col=sheet.max_column):
                for cell in row:

                    if is_comment and cell.comment is not None:
                        comment = cell.comment.text.split("\n")[1]
                        if comment not in ["total_time", "csat"]:
                            result[name].update({comment: cell.column})

                    if not is_comment and cell.comment is None and cell.value is not None:
                        result[name].append(cell.column)
        return result

    @staticmethod
    def get_column_time(sheet: Worksheet):
        for row in sheet.iter_rows(max_row=1, min_col=2, max_col=sheet.max_column):
            for cell in row:
                if cell.comment is not None and cell.comment.text.split("\n")[1] == "total_time":
                    return cell.column

    @staticmethod
    def get_column_csat(sheet: Worksheet):
        for row in sheet.iter_rows(max_row=1, min_col=2, max_col=sheet.max_column):
            for cell in row:
                if cell.comment is not None and cell.comment.text.split("\n")[1] == "csat":
                    return cell.column

    @staticmethod
    def add_comment(sheet: Worksheet, row, date):
        cell = sheet.cell(row=row, column=1, value=date)
        cell.comment = Comment(f"{datetime.now().strftime('%H:%M:%S')}", "auto")
        sheet[f"A{row + 1}"].number_format = "DD.MM.YYYY"

    def do_offset_formulas(self, sheet: Worksheet, data: list, row: int):
        for col in data:

            try:
                cell: Cell = sheet.cell(row=row, column=col)
                cell_offset = cell.offset(row=-1)

                if isinstance(cell_offset.value, str) and cell_offset.value.startswith("="):
                    sheet[cell.coordinate] = cell_offset.value.replace(str(cell_offset.row), str(row))

                    if "РСД" not in cell_offset.value:
                        sheet[cell.coordinate].number_format = BUILTIN_FORMATS[9]

            except Exception:
                print("Ошибка в записи формул!")
                self.logger.error(traceback.format_exc())

    def set_total_time(self, sheet: Worksheet, date, row):
        column = self.get_column_time(sheet)
        if column is not None:
            time = self.api.get_total_time(date)
            sheet.cell(row=row, column=column, value=time)

    def set_csat(self, sheet: Worksheet, date, row):
        column = self.get_column_csat(sheet)
        if column is not None:
            csat = self.api.get_csat(date)

            summa = sum(csat.values())
            value = (csat.get('5', 0) + csat.get('4', 0)) / summa if summa != 0 else 0

            cell: Cell = sheet.cell(row=row, column=column, value=value)
            sheet[cell.coordinate].number_format = BUILTIN_FORMATS[10]

    def main(self):
        book = openpyxl.load_workbook(self.filename)

        goals = self.get_ids_refinancing(book)
        formulas = self.get_ids_refinancing(book, is_comment=False)

        for name, ids in goals.items():
            sheet: Worksheet = book[name]
            last_row = self.get_row_for_write(sheet)
            first_date: datetime = last_row[0]

            if first_date.date() <= datetime.now().date():
                row = last_row[1]

                for date in pd.date_range(first_date, self.get_now.strftime('%Y-%m-%d')):
                    # добавляет коментарий с датой выгрузки
                    self.add_comment(sheet, row, date)
                    # копирует формулы с предыдущих ячеек
                    self.do_offset_formulas(sheet, formulas[name], row)

                    date_format = date.strftime("%Y-%m-%d")
                    # устанавливает общее время
                    self.set_total_time(sheet, date_format, row)
                    # устанавливает csat
                    self.set_csat(sheet, date_format, row)
                    metrics = api_yandex_async.main(ids, date1=date_format, date2=date_format, need_users=False)

                    self.logger.info(f"{name}: {metrics}")

                    # записывает полученные метрики
                    for col, goal in metrics.items():
                        if goal != "" and col != "date" and col != "total_time":
                            sheet.cell(row=row, column=col, value=int(goal))

                    row += 1

        book.save(self.filename)
        book.close()

        plyer.notification.notify(message='Metrics is download!',
                                  app_name='ParsMetrics',
                                  app_icon='data/static/success_icon-icons.com_52365.ico',
                                  title='Success',
                                  timeout=2)
        
        next_run = schedule.next_run().strftime("%H:%M:%S")
        self.logger.info(f"Выгрузка завершена. Следующая в {next_run}")

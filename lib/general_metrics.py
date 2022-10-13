import json

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from common.editor_excel import EditorExcel
from common import BaseExcel, api_yandex_async


class GeneralMetrics(BaseExcel):
    def __init__(self, filename=None):
        super().__init__()
        self.filename = self.get_filepath("Метрики2022КОПИЯ") if filename is None else self.get_filepath(filename)
        self.book = openpyxl.load_workbook(self.filename)

    # def get_ids_general(self, sheet: Worksheet) -> dict[dict] | dict[list]:
    #     result = {}
    #     for column in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=1):
    #         for cell in column:
    #             if cell.comment is not None:
    #                 comment = cell.comment.text.split("\n")[1]
    #                 result[column].update({comment: cell.column})
    #     return result

    def name_sheet(self, month=None):
        if month is None:
            month = self.get_yesterday.month
        return json.load(open(r"data/name_sheet.json", encoding="utf-8"))[month]  # текущий месяц

    def main(self, day=None, month=None, date1='yesterday', date2='yesterday'):
        if day is None:
            day = str(self.get_yesterday.day)

        if month is None:
            month = str(self.get_yesterday.month)

        ids: dict = EditorExcel(self.filename).get_id_row("BR")
        metrics = api_yandex_async.main(ids, date1=date1, date2=date2)  # выгруженные метрики

        sheet = self.book[self.name_sheet(month)]  # нужный лист в ecxel
        col = json.load(open(r'data/date_col.json', encoding="utf-8"))[day]  # нужная колонка

        for row, goal in metrics.items():
            if row != "date" and goal != '':
                sheet[col + str(row)] = int(goal)  # запись значения в ячейки

        self.book.save(filename=self.filename)  # сохранение изменений
